import openpyxl
from openpyxl import workbook

def generatePreferences(values):

    # make an empty dictionary for the preference profile
    preferences = {}
    
    # Get the number of rows (agents) and columns (alternatives) from the values worksheet
    num_rows = values.max_row
    num_cols = values.max_column
    
    # Iterate over the rows (agents)
    for i in range(1, num_rows+1):
        # make an empty list for the current agent's preference ordering
        pref_ordering = []
        # Get the valuations of the current agent for all alternatives
        valuations = [values.cell(row=i, column=j).value for j in range(1, num_cols+1)]
        # Sort the alternatives based on the valuations
        sorted_alternatives = [alt for _, alt in sorted(zip(valuations, range(1, num_cols+1)), reverse=True)]
        # Add the sorted alternatives to the current agent's preference ordering
        pref_ordering.extend(sorted_alternatives)
        # Add the preference ordering of the current agent to the preference profile
        preferences[i] = pref_ordering
    
    return preferences




def dictatorship(preferenceProfile, agent):
    """
    Decides the winner in an election by the dictatorship method.
    
    Parameters:
    preferenceProfile (dict): A dictionary containing the preference ordering of each agent. The keys are the indices of the agents
    (starting from 1) and the values are lists containing the indices of the alternatives (also starting from 1) ordered
    according to the agents' preferences.
    agent (int): The index of the agent whose preference will be used to determine the winner (starting from 1).
    
    Returns:
    int: The index of the winning alternative (starting from 1).
    
    Example:
    preferenceProfile = {
        1: [2, 3, 1],  # Agent 1 prefers alternative 2, then 3, then 1
        2: [1, 2, 3],  # Agent 2 prefers alternative 1, then 2, then 3
        3: [3, 1, 2],  # Agent 3 prefers alternative 3, then 1, then 2
    }
    dictatorship(preferenceProfile, 1)  # Returns: 2 (alternative 2 is the first choice of agent 1)
    """
    #checks if specified agent present in the preference profile 
    if agent not in preferenceProfile:
        raise ValueError("Agent does not exist")
    #returns the first choice if specified agent
    return preferenceProfile[agent][0]


#For some functions tie breaker function is called outside from it and for some function tie break rule is implied in it

def scoringRule(preferences, scoreVector, tieBreak):
    num_agents = len(preferences.keys())
    num_alternatives = len(preferences[1])
    try:
        assert len(scoreVector) == num_alternatives
    except:
        print('Incorrect input')
        return False
    else:
        scoreVector = sorted(scoreVector, reverse=True)
        prefList = list(preferences.values())
        totalScore = {}
        for I in prefList:
            for i, k in enumerate(I):
                totalScore[k] = totalScore.get(k, 0) + scoreVector[i]
        all_values = totalScore.values()
        max_value = max(all_values)
        winners = [k for k, v in totalScore.items() if v == max_value]
    if tieBreak == "max": 
        return max(winners)
    elif tieBreak == "min":
        return min(winners)
    else:
        # Check if tieBreak corresponds to a valid agent
      try:
        agent_i = int(tieBreak)
        assert 1 <= agent_i <= num_agents

            # Get the preference ordering of the specified agent
        agentPref = preferences[agent_i]
            
            # Find the index of the first occurrence of each of the winning alternatives in the agent's preference ordering
        winner_indexes = [agentPref.index(i) for i in winners]
            
            # Find the index of the minimum of the winner indexes
        winner_index = min(winner_indexes)
            
            # Get the winning alternative at the winner index
        winner_agent = agentPref[winner_index]
        return winner_agent
            
      except:
        print("this input does not correspond to any agent.")


def plurality(preferences, tieBreak) -> int:
    frequency = {}
    frequency_list = [0] * (len(preferences[1]) + 1)
    for key, row in preferences.items():
        frequency_list[row[0]] += 1
    countList = []
    for index in range(len(frequency_list)):
        if frequency_list[index] in frequency.keys():
            countList = list.copy(frequency[frequency_list[index]])
        if (frequency_list[index] != 0):
            countList.append(index)
            frequency[frequency_list[index]] = list.copy(countList)
            countList.clear()
    maxOccurence = max(frequency.keys())
    if len(frequency[maxOccurence]) == 1:
        return frequency[maxOccurence][0]
    else:
        return tieBreakHelper(tieBreak, frequency[maxOccurence], preferences)

 
def veto(preferences, tieBreak):
    alternatives = list(set([a for preference in preferences.values() for a in preference]))
    n = len(preferences)  
    points = {a: 0 for a in alternatives}
    for agent, preference in preferences.items():
        for i, a in enumerate(preference):
            if i != len(preference) - 1:
                points[a] += 1
    max_points = max(points.values())
    winners = [a for a, p in points.items() if p == max_points]
    if tieBreak == "max":
        return max(winners, key=alternatives.index)
    elif tieBreak == "min":
        return min(winners, key=alternatives.index) 
    elif isinstance(tieBreak, int) and 1 <= tieBreak <= n:
        return sorted(winners, key=lambda x: preferences[tieBreak].index(x))[0]
    else:
        raise ValueError("Invalid tie-breaking rule")





def borda(preference_profile, tieBreak) -> int:
  """
  Calculate the winner of an election based on a Borda count of the preferences of each agent.
  
  Args:
    preference_profile (dict): A dictionary where the keys are the agents (strings) and the values 
        are the preferences of that agent (list of strings).
    tieBreak (function): A function that takes in a list of candidates (strings) and a preference 
        profile (dict) and returns the winning candidate (string). This function is used to break 
        ties when multiple candidates have the same lowest score.
  
  Returns:
    The winning candidate (string). If there is a tie, the tieBreak function is used to determine 
        the winner. If no preferences are provided or all preferences are the same, returns None.
  """
  scores = {}
  # Count the number of times each candidate is ranked and store the counts in a dictionary.
  for agent, preferences in preference_profile.items():
    for i, preference in enumerate(preferences):
      if preference not in scores:
        scores[preference] = 0
      scores[preference] += i
  if not scores:
    return None
  # Find the minimum score and the candidates with the minimum score.
  min_score = min(scores.values())
  if min_score == 0:
    return None
  winners = [a for a, s in scores.items() if s == min_score]
  # If there is only one winner, return it.
  if len(winners) == 1:
    return winners[0]
  # If there is a tie, use the tieBreak function to determine the winner.
  return tieBreakHelper(tieBreak, winners, preference_profile)





def harmonic(preferences, tie_break):
    scores = {}
    # Count the number of times each candidate is ranked and store the harmonic mean in a dictionary.
    for agent, ranking in preferences.items():
        for i, alt in enumerate(ranking):
            if alt not in scores:
                scores[alt] = 0
            scores[alt] += 1 / (i + 1)
    # Find the maximum score and the candidates with the maximum score.
    max_score = max(scores.values())
    winners = [a for a, s in scores.items() if s == max_score]
    # If there is only one winner, return it.
    if len(winners) == 1:
        return winners[0]
    else:
        # If the tie_break argument is "max", return the  maximum candidate.
        if tie_break == "max":
            return max(winners)
        # If the tie_break argument is "min", return the  minimum candidate.
        elif tie_break == "min":
            return min(winners)
        # If the tie_break argument is an integer, return the first preference of the agent with 
        # that integer as the key.
        else:
            try:
                i = int(tie_break)
                return preferences[i][0]
            # If the tie_break argument is invalid, return False.
            except (ValueError, KeyError):
                return False


import copy 

def STV(preferences, tieBreak):
    """
    This function implements the Single Transferable Vote (STV) method for conducting an election. The STV method is a type of ranked voting system in which voters rank the candidates in order of preference.

    Parameters:
    - preferences (dict): a dictionary where the keys are the voter IDs and the values are lists representing the rankings of the candidates by that voter
    - tieBreak (dict): a dictionary that provides a tie-breaking rule for resolving ties in the election. The keys are the candidate IDs and the values are the tie-breaking values for those candidates.

    Returns:
    - int: the ID of the winning candidate
    """
    tempPreferences = copy.deepcopy(list(preferences.values()))
    deletedSet = set()
    frequency_list = [0] * (len(preferences[1]) + 1)
    size = len(tempPreferences[0])
    while (size > 0):
        # Count the number of votes for each candidate
        for row in tempPreferences:
            frequency_list[row[0]] += 1

        # Create a dictionary that maps the frequency of votes for each candidate to a list of candidates with that frequency
        frequency = {}
        countList = []
        for index in range(len(frequency_list)):
            if frequency_list[index] in frequency.keys():
                countList = list.copy(frequency[frequency_list[index]])
            if (index != 0):
                countList.append(index)
                frequency[frequency_list[index]] = list.copy(countList)
                countList.clear()

        # Determine the minimum frequency of votes among the candidates
        minimumkey = min(frequency.keys())

        # If there are two or fewer candidates with the minimum frequency, resolve the tie using the tie-breaking rule
        if (len(frequency.keys()) <= 2):
            winningValue = tieBreakHelper(tieBreak, frequency[minimumkey], preferences)
            return winningValue
        else:
            # Remove all candidates with the minimum frequency of votes from tempPreferences and deletedSet, and set their
            # frequency in frequency_list to 999 to indicate that they have been eliminated
            for val in frequency[minimumkey]:
                if (val not in deletedSet):
                    for values in tempPreferences:
                        values.remove(val)
                        deletedSet.add(val)
                        frequency_list[val] = 999

        # Reset all elements in frequency_list to 0, except those that have been set to 999
        frequency_list = [999 if row == 999 else 0 for row in frequency_list]
    return 0




def tieBreakHelper(tiebreak, alternatives, preferenceProfile) -> int:
    # If the tiebreak argument is "max", return the  maximum candidate.
    if tiebreak == "max":
        return max(alternatives)
    # If the tiebreak argument is "min", return the  minimum candidate.
    elif tiebreak == "min":
        return min(alternatives)
    else:
        # If the tiebreak argument is a string representation of an integer, return the first 
        # preference of the agent with that integer as the key.
        try:
            agentValue = int(tiebreak)
            assert agentValue in preferenceProfile.keys()
            maxRanked = 9999999
            maxNumber = -1
            for row in alternatives:
                if (preferenceProfile[agentValue].index(row) < maxRanked):
                    maxRanked = preferenceProfile[agentValue].index(row)
                    maxNumber = row
            return maxNumber
        # If the tiebreak argument is invalid, print an error message and return None.
        except AssertionError:
            print("Not a valid agent, please enter a valid agent")
        except ValueError:
            print("Entered row is not a digit")






def rangeVoting(values, tieBreak):
    # Generate a preference profile from the values DataFrame.
    agentdict = generatePreferences(values)
    # Count the total score for each candidate.
    frequency = {}
    frequency_list = [0] * (len(list(agentdict.values())[0]) + 1)
    index = 0
    for value in agentdict.values():
        numericalValues = [row.value for row in list(values.rows)[index]]
        index += 1
        for v in value:
            frequency_list[v] += float(numericalValues[v - 1])
    # Find the candidates with the maximum total score.
    countList = []
    for index in range(len(frequency_list)):
        if frequency_list[index] in frequency.keys():
            countList = list.copy(frequency[frequency_list[index]])
        if (frequency_list[index] != 0):
            countList.append(index)
            frequency[frequency_list[index]] = list.copy(countList)
            countList.clear()
    maxOccurence = max(frequency.keys())
    # If there is only one candidate with the maximum total score, return it.
    if len(frequency[maxOccurence]) == 1:
        return frequency[maxOccurence][0]
    # If there is a tie, use the tieBreak function to determine the winner.
    else:
        return tieBreakHelper(tieBreak, frequency[maxOccurence], agentdict)
